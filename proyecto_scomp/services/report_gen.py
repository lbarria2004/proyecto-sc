from fpdf import FPDF
import io
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from utils.helpers import clean_number

class PDFReportVejez(FPDF):
    
    def __init__(self, orientation='L', unit='mm', format='A4'):
        super().__init__(orientation, unit, format)
        self.left_margin = 15
        self.right_margin = 15
        self.top_margin = 15
        self.set_auto_page_break(auto=True, margin=15)
        
    def add_page(self, orientation='', format='', same=False, rotation=0):
        super().add_page(orientation=orientation, format=format)
        self.set_left_margin(self.left_margin)
        self.set_right_margin(self.right_margin)
        self.set_top_margin(self.top_margin)

    def header(self):
        self.set_font('Arial', 'B', 18)
        self.cell(0, 10, 'ESTUDIO PRELIMINAR DE PENSIÓN', 0, 1, 'C')
        self.ln(5)

    def print_header_data(self, data):
        row_height = 7
        self.set_font('Arial', 'B', 11)
        self.cell(40, row_height, "Afiliado:", 0, 0)
        self.set_font('Arial', '', 11)
        self.cell(0, row_height, data.get('nombre', 'No encontrado').encode('latin-1', 'replace').decode('latin-1'), 0, 1)
        
        self.set_font('Arial', 'B', 11)
        self.cell(40, row_height, "RUT:", 0, 0)
        self.set_font('Arial', '', 11)
        self.cell(0, row_height, str(data.get('rut', 'No encontrado')), 0, 1)
        
        self.set_font('Arial', 'B', 11)
        self.cell(40, row_height, "Saldo Acumulado:", 0, 0)
        self.set_font('Arial', '', 11)
        self.cell(0, row_height, str(data.get('saldo_uf', 'No encontrado')), 0, 1)
        
        self.set_font('Arial', 'B', 11)
        self.cell(40, row_height, "N° SCOMP:", 0, 0)
        self.set_font('Arial', '', 11)
        self.cell(0, row_height, str(data.get('n_scomp', 'No encontrado')), 0, 1)
        
        self.set_font('Arial', 'B', 11)
        self.cell(40, row_height, "Tipo Pensión:", 0, 0)
        self.set_font('Arial', '', 11)
        self.cell(0, row_height, data.get('tipo_pension', 'No encontrado').encode('latin-1', 'replace').decode('latin-1'), 0, 1)

        self.set_font('Arial', 'B', 11)
        self.cell(40, row_height, "Valor UF:", 0, 0)
        self.set_font('Arial', '', 11)
        self.cell(0, row_height, str(data.get('valor_uf_str', 'No encontrado')), 0, 1)

        if data.get('pgu_monto'):
             self.set_font('Arial', 'B', 11)
             self.cell(40, row_height, "Monto PGU:", 0, 0)
             self.set_font('Arial', '', 11)
             self.cell(0, row_height, f"${data.get('pgu_monto'):,.0f}", 0, 1)
        
        if data.get('bono_monto_uf'):
             self.set_font('Arial', 'B', 11)
             self.cell(40, row_height, "Bono x Años:", 0, 0)
             self.set_font('Arial', '', 11)
             self.cell(0, row_height, f"{data.get('bono_monto_uf')} UF", 0, 1)

        self.ln(8)

    def print_table(self, title, df, tipo, col_title_pdf, eld_info, valor_uf=0.0):
        safe_title = title.encode('latin-1', 'replace').decode('latin-1')
        
        self.set_font('Arial', 'B', 12)
        self.set_fill_color(230, 230, 230)
        self.cell(0, 8, safe_title, 1, 1, 'C', fill=True)
        
        self.set_font('Arial', 'B', 9)
        row_height = 7
        usable_width = self.w - self.l_margin - self.r_margin 
        
        if tipo == 'RP':
            widths = {
                'modalidad': usable_width * 0.22, 'uf': usable_width * 0.12, 'bruta': usable_width * 0.14,
                'salud': usable_width * 0.14, 'comision': usable_width * 0.14, 'liquida': usable_width * 0.14,
                'adicional': usable_width * 0.10
            }
            self.cell(widths['modalidad'], row_height, "Modalidad", 1, 0, 'C')
            self.cell(widths['uf'], row_height, "Pensión UF", 1, 0, 'C')
            self.cell(widths['bruta'], row_height, "Pensión Bruta", 1, 0, 'C')
            self.cell(widths['salud'], row_height, "Dscto. 7% Salud", 1, 0, 'C')
            self.cell(widths['comision'], row_height, "Dscto. Comisión", 1, 0, 'C')
            self.cell(widths['liquida'], row_height, "Pensión Líquida", 1, 0, 'C')
            self.cell(widths['adicional'], row_height, col_title_pdf, 1, 1, 'C')
            
        elif tipo == 'RT': # Renta Temporal
            widths = {
                'compania': usable_width * 0.22, 'uf': usable_width * 0.12, 'bruta': usable_width * 0.14,
                'salud': usable_width * 0.14, 'comision': usable_width * 0.14, 'liquida': usable_width * 0.14,
                'adicional': usable_width * 0.10
            }
            self.cell(widths['compania'], row_height, "Compañía", 1, 0, 'C')
            self.cell(widths['uf'], row_height, "Pensión UF", 1, 0, 'C')
            self.cell(widths['bruta'], row_height, "Pensión Bruta", 1, 0, 'C')
            self.cell(widths['salud'], row_height, "Dscto. 7% Salud", 1, 0, 'C')
            self.cell(widths['comision'], row_height, "Dscto. Comisión", 1, 0, 'C')
            self.cell(widths['liquida'], row_height, "Pensión Líquida", 1, 0, 'C')
            self.cell(widths['adicional'], row_height, col_title_pdf, 1, 1, 'C')

        else: # RV, RVD, RV_Aumentada, RV_Base, REF
            widths = {
                'compania': usable_width * 0.28, 'uf': usable_width * 0.12, 'bruta': usable_width * 0.15,
                'salud': usable_width * 0.15, 'liquida': usable_width * 0.15, 'adicional': usable_width * 0.15
            }
            self.cell(widths['compania'], row_height, "Compañía", 1, 0, 'C')
            self.cell(widths['uf'], row_height, "Pensión UF", 1, 0, 'C')
            self.cell(widths['bruta'], row_height, "Pensión Bruta", 1, 0, 'C')
            self.cell(widths['salud'], row_height, "Dscto. 7% Salud", 1, 0, 'C')
            self.cell(widths['liquida'], row_height, "Pensión Líquida", 1, 0, 'C')
            self.cell(widths['adicional'], row_height, col_title_pdf, 1, 1, 'C')

        self.set_font('Arial', '', 9)
        for _, row in df.iterrows():
            compania_safe = str(row.get('Compania', 'N/A')).encode('latin-1', 'replace').decode('latin-1')
            modalidad_safe = str(row.get('Modalidad', 'N/A')).encode('latin-1', 'replace').decode('latin-1')
            pension_uf_str = str(row.get('Pension_UF', 'N/A'))
            
            pension_bruta_str = f"${row['Pension_Bruta']:,.0f}"
            descuento_salud_str = f"(${row['Descuento_Salud_7%']:,.0f})"
            pension_liquida_str = f"${row['Pension_Liquida']:,.0f}"
            pension_extra_str = f"${row[col_title_pdf]:,.0f}"

            if tipo == 'RP':
                comision_val = row.get('Descuento_Comision', 0)
                comision_str = f"(${comision_val:,.0f})" if comision_val > 0 else "($0)"
                
                self.cell(widths['modalidad'], row_height, modalidad_safe, 1, 0)
                self.cell(widths['uf'], row_height, pension_uf_str, 1, 0, 'R')
                self.cell(widths['bruta'], row_height, pension_bruta_str, 1, 0, 'R')
                self.cell(widths['salud'], row_height, descuento_salud_str, 1, 0, 'R')
                self.cell(widths['comision'], row_height, comision_str, 1, 0, 'R')
                self.cell(widths['liquida'], row_height, pension_liquida_str, 1, 0, 'R')
                self.cell(widths['adicional'], row_height, pension_extra_str, 1, 1, 'R')
            
            elif tipo == 'RT': # Renta Temporal
                comision_val = row.get('Descuento_Comision', 0)
                comision_str = f"(${comision_val:,.0f})" if comision_val > 0 else "($0)"
                
                self.cell(widths['compania'], row_height, compania_safe, 1, 0)
                self.cell(widths['uf'], row_height, pension_uf_str, 1, 0, 'R')
                self.cell(widths['bruta'], row_height, pension_bruta_str, 1, 0, 'R')
                self.cell(widths['salud'], row_height, descuento_salud_str, 1, 0, 'R')
                self.cell(widths['comision'], row_height, comision_str, 1, 0, 'R')
                self.cell(widths['liquida'], row_height, pension_liquida_str, 1, 0, 'R')
                self.cell(widths['adicional'], row_height, pension_extra_str, 1, 1, 'R')

            else: # RV, RVD, etc.
                self.cell(widths['compania'], row_height, compania_safe, 1, 0)
                self.cell(widths['uf'], row_height, pension_uf_str, 1, 0, 'R')
                self.cell(widths['bruta'], row_height, pension_bruta_str, 1, 0, 'R')
                self.cell(widths['salud'], row_height, descuento_salud_str, 1, 0, 'R')
                self.cell(widths['liquida'], row_height, pension_liquida_str, 1, 0, 'R')
                self.cell(widths['adicional'], row_height, pension_extra_str, 1, 1, 'R')
        
        if eld_info and eld_info.get("monto_pesos") is not None and eld_info.get("monto_pesos", 0) > 0:
            self.set_font('Arial', 'B', 10) 
            self.ln(2)
            
            note_text = ""
            if tipo == 'RP':
                monto_uf = eld_info.get("monto_uf", "N/A")
                monto_pesos_str = f"${eld_info.get('monto_pesos', 0):,.0f}"
                p_uf = eld_info.get("pension_resultante_uf", "N/A")
                
                p_uf_float = clean_number(p_uf)
                p_pesos_calculados = p_uf_float * valor_uf
                p_pesos_str = f"${p_pesos_calculados:,.0f}"
                
                note_text = (
                    f"Opcion de Excedente (ELD) en Retiro Programado:\n"
                    f"Monto ELD: {monto_uf} UF ({monto_pesos_str}) con Pension resultante: {p_uf} UF ({p_pesos_str})"
                )
            
            elif tipo in ['RV', 'RV_Aumentada', 'REF', 'RT']:
                compania = eld_info.get("compania", "N/A")
                monto_str = f"${eld_info.get('monto_pesos', 0):,.0f}"
                compania_safe = compania.encode('latin-1', 'replace').decode('latin-1')
                note_text = f"Bajo esta modalidad, la compania que ofrece el mayor monto de Excedente de Libre Disposicion es {compania_safe} con {monto_str} pesos."
            
            self.multi_cell(0, 5, note_text.encode('latin-1', 'replace').decode('latin-1'), 0, 'L')
            self.ln(5) 
            
        else:
            if tipo not in ['RV_Aumentada', 'RT']: 
                self.ln(10)

class PDFReportSobrevivencia(FPDF):
    
    def __init__(self, orientation='L', unit='mm', format='A4'): 
        super().__init__(orientation, unit, format)
        self.header_data = {}
        self.beneficiarios_ordenados = []
        self.l_margin = 15 
        self.r_margin = 15
        self.t_margin = 15
        self.set_auto_page_break(auto=True, margin=15)

    def set_header_data(self, data, beneficiarios):
        self.header_data = data
        self.beneficiarios_ordenados = beneficiarios

    def header(self):
        self.set_font('Arial', 'B', 16)
        self.cell(0, 10, 'RESULTADO SCOMP', 0, 1, 'C')
        self.set_font('Arial', 'B', 12)
        
        nombre_consultante = self.header_data.get('nombre', 'N/A').encode('latin-1', 'replace').decode('latin-1')
        self.cell(0, 7, f"SRA. {nombre_consultante}", 0, 1, 'L')
        
        valor_uf = self.header_data.get('valor_uf_str', 'N/A')
        self.cell(0, 7, f"VALOR UF: {valor_uf}", 0, 1, 'L')
        self.ln(5)

    def print_header_data_sobrevivencia(self):
        self.set_font('Arial', 'B', 11)
        self.cell(40, 7, "Consultante:", 0, 0)
        self.set_font('Arial', '', 11)
        self.cell(0, 7, self.header_data.get('nombre', 'N/A').encode('latin-1', 'replace').decode('latin-1'), 0, 1)
        
        self.set_font('Arial', 'B', 11)
        self.cell(40, 7, "RUT:", 0, 0)
        self.set_font('Arial', '', 11)
        self.cell(0, 7, str(self.header_data.get('rut', 'N/A')), 0, 1)
        
        self.set_font('Arial', 'B', 11)
        self.cell(40, 7, "Saldo Acumulado:", 0, 0)
        self.set_font('Arial', '', 11)
        self.cell(0, 7, str(self.header_data.get('saldo_uf', 'No encontrado')), 0, 1)
        
        self.set_font('Arial', 'B', 11)
        self.cell(40, 7, "N° SCOMP:", 0, 0)
        self.set_font('Arial', '', 11)
        self.cell(0, 7, str(self.header_data.get('n_scomp', 'No encontrado')), 0, 1)
        
        self.set_font('Arial', 'B', 11)
        self.cell(40, 7, "Tipo Pensión:", 0, 0)
        self.set_font('Arial', '', 11)
        self.cell(0, 7, self.header_data.get('tipo_pension', 'No encontrado').encode('latin-1', 'replace').decode('latin-1'), 0, 1)

        # Imprimir Beneficiarios
        self.ln(5)
        self.set_font('Arial', 'B', 12)
        self.cell(0, 7, "Beneficiarios", 0, 1, 'L')
        self.set_font('Arial', '', 10)
        
        for i, b in enumerate(self.beneficiarios_ordenados):
            line = f"{i + 1}) {b['nombre']} ({b['parentesco']})"
            self.cell(0, 6, line.encode('latin-1', 'replace').decode('latin-1'), 0, 1)
        self.ln(8)

    def print_table_sobrevivencia(self, title, df, tipo):
        safe_title = title.encode('latin-1', 'replace').decode('latin-1')
        
        self.set_font('Arial', 'B', 12)
        self.set_fill_color(230, 230, 230)
        self.cell(0, 8, safe_title, 1, 1, 'C', fill=True)
        
        self.set_font('Arial', 'B', 8) 
        row_height = 7
        usable_width = self.w - self.l_margin - self.r_margin 
        
        if tipo == 'RP_SOBREVIVENCIA':
            cols = ["Beneficiario", "Pension_UF", "Pension_Bruta", "Dscto_Salud_7%", "Dscto_Comision", "Pension_Liquida"]
            widths = [usable_width * 0.30] + [usable_width * 0.14] * 5 
            
            for i, col_name in enumerate(cols):
                self.cell(widths[i], row_height, col_name, 1, 0, 'C')
            self.ln()
            
            self.set_font('Arial', '', 8)
            for _, row in df.iterrows():
                is_total = (row['Beneficiario'] == 'Pensión mensual total')
                if is_total:
                    self.set_font('Arial', 'B', 8) 
                
                self.cell(widths[0], row_height, str(row['Beneficiario']).encode('latin-1', 'replace').decode('latin-1'), 1, 0, 'L')
                self.cell(widths[1], row_height, str(row['Pension_UF']), 1, 0, 'R')
                self.cell(widths[2], row_height, f"${row['Pension_Bruta']:,.0f}", 1, 0, 'R')
                self.cell(widths[3], row_height, f"(${row['Dscto_Salud_7%']:,.0f})", 1, 0, 'R')
                self.cell(widths[4], row_height, f"(${row['Dscto_Comision']:,.0f})", 1, 0, 'R')
                self.cell(widths[5], row_height, f"${row['Pension_Liquida']:,.0f}", 1, 1, 'R')
                
                if is_total:
                    self.set_font('Arial', '', 8)

        elif tipo == 'RV_SOBREVIVENCIA':
            cols = ["Compañía"]
            num_benef = len(self.beneficiarios_ordenados)
            for i in range(num_benef):
                cols.extend([f"Benef. {i+1} UF", f"Benef. {i+1} $"])
            cols.extend(["Total Bruto", "Total Líquido"])

            num_benef_cols = num_benef * 2
            
            width_compania = usable_width * 0.20
            width_totales = usable_width * 0.12
            width_benef_cols = (usable_width - width_compania - (width_totales * 2)) / num_benef_cols if num_benef_cols > 0 else 0
            
            widths = [width_compania]
            for i in range(num_benef_cols):
                widths.append(width_benef_cols)
            widths.extend([width_totales, width_totales])
            
            # Imprimir cabeceras
            for i, col_name in enumerate(cols):
                self.cell(widths[i], row_height, col_name.encode('latin-1', 'replace').decode('latin-1'), 1, 0, 'C')
            self.ln()
            
            # Imprimir filas
            self.set_font('Arial', '', 8)
            for _, row in df.iterrows():
                for i, col_name in enumerate(df.columns): 
                    align = 'L' if i == 0 else 'R'
                    cell_value = row[col_name]
                    
                    if isinstance(cell_value, (int, float)):
                        cell_text = f"${cell_value:,.0f}"
                    else:
                        cell_text = str(cell_value).encode('latin-1', 'replace').decode('latin-1')
                        
                    self.cell(widths[i], row_height, cell_text, 1, 0, align)
                self.ln()
        
        self.ln(10) 

def create_formatted_excel_report(header_data, processed_tables, beneficiarios=None):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte SCOMP"

    # --- Definir Estilos ---
    style_header_title = Font(name='Arial', size=14, bold=True)
    style_table_title_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    style_table_title_font = Font(name='Arial', size=12, bold=True)
    style_table_title_align = Alignment(horizontal='center', vertical='center')
    style_table_header_font = Font(name='Arial', size=10, bold=True)
    style_table_header_align = Alignment(horizontal='center')
    style_bold_text = Font(name='Arial', size=10, bold=True)
    
    thin_border_side = Side(border_style="thin", color="000000")
    style_table_border = Border(top=thin_border_side, left=thin_border_side, right=thin_border_side, bottom=thin_border_side)

    # Estilo de moneda CON bordes
    style_currency = NamedStyle(name='currency_style', number_format='$#,##0') 
    style_currency.border = style_table_border
    if 'currency_style' not in wb.style_names:
        wb.add_named_style(style_currency)
        
    # Estilo UF CON bordes
    style_uf = NamedStyle(name='uf_style', number_format='0.00')
    style_uf.border = style_table_border
    if 'uf_style' not in wb.style_names:
        wb.add_named_style(style_uf)
    
    current_row = 1

    # --- 1. Escribir Header Data ---
    ws[f'A{current_row}'] = "Resultado Resumen de Scomp"
    ws[f'A{current_row}'].font = style_header_title
    current_row += 2
    
    header_map = {
        "nombre": "Consultante/Afiliado",
        "rut": "RUT",
        "tipo_pension": "Tipo Pensión",
        "saldo_uf": "Saldo Acumulado",
        "n_scomp": "N° SCOMP",
        "valor_uf_str": "Valor UF"
    }
    
    for key, label in header_map.items():
        ws[f'A{current_row}'] = label
        ws[f'A{current_row}'].font = style_bold_text
        ws[f'B{current_row}'] = header_data.get(key, 'N/A')
        current_row += 1
    
    current_row += 1 

    # --- 2. Escribir Beneficiarios (si existen) ---
    if beneficiarios:
        ws[f'A{current_row}'] = "Beneficiarios"
        ws[f'A{current_row}'].font = style_header_title
        current_row += 1
        
        headers = ["ID", "Nombre", "RUT", "Parentesco"]
        for c_idx, header_title in enumerate(headers):
            cell = ws.cell(row=current_row, column=c_idx + 1, value=header_title)
            cell.font = style_table_header_font
            cell.border = style_table_border
            cell.alignment = style_table_header_align
        current_row += 1

        for i, b in enumerate(beneficiarios):
            cellA = ws.cell(row=current_row, column=1, value=f"Benef. {i+1}")
            cellB = ws.cell(row=current_row, column=2, value=b.get('nombre'))
            cellC = ws.cell(row=current_row, column=3, value=b.get('rut'))
            cellD = ws.cell(row=current_row, column=4, value=b.get('parentesco'))
            
            for cell in [cellA, cellB, cellC, cellD]:
                cell.border = style_table_border
            current_row += 1
        current_row += 1 

    # --- 3. Escribir cada tabla de modalidad ---
    column_widths = {}
    
    for item in processed_tables:
        df_table = item['tabla']
        title = item['titulo']
        
        if 'Comision_Pct' in df_table.columns:
            df_table = df_table.drop(columns=['Comision_Pct'])
        
        # Escribir Título de la tabla y Combinar Celdas
        num_cols = len(df_table.columns)
        if item.get('tipo') == 'RP_SOBREVIVENCIA':
             num_cols += 1
            
        ws[f'A{current_row}'] = title
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
        cell_title = ws[f'A{current_row}']
        cell_title.fill = style_table_title_fill
        cell_title.font = style_table_title_font
        cell_title.alignment = style_table_title_align
        current_row += 1
        
        # Escribir filas de la tabla (con cabeceras)
        include_index = (item.get('tipo') == 'RP_SOBREVIVENCIA')
        
        rows = list(dataframe_to_rows(df_table, index=include_index, header=True))
        
        header_names = rows[0]

        for r_idx, row in enumerate(rows):
            for c_idx, value in enumerate(row):
                cell = ws.cell(row=current_row + r_idx, column=c_idx + 1, value=value)
                
                cell.border = style_table_border
                
                col_letter = cell.column_letter
                current_width = column_widths.get(col_letter, 0)
                
                value_str = ""
                if isinstance(value, (int, float)):
                    if "UF" in (header_names[c_idx] or ""):
                        value_str = f"{value:.2f}"
                    else:
                        value_str = f"${value:,.0f}"
                else:
                    value_str = str(value)
                
                value_length = len(value_str)
                header_length = len(str(header_names[c_idx]))
                
                column_widths[col_letter] = max(current_width, value_length, header_length)
                
                if r_idx == 0: # Cabecera
                    cell.font = style_table_header_font
                    cell.alignment = style_table_header_align
                else: # Datos
                    if isinstance(value, (int, float)):
                        col_name = header_names[c_idx] if c_idx < len(header_names) else ""
                        if col_name is None: col_name = ""
                        
                        if "$" in col_name or "Bruta" in col_name or "Líquida" in col_name or "Dscto" in col_name or "Total" in col_name:
                            cell.style = 'currency_style'
                        elif "UF" in col_name:
                            cell.style = 'uf_style'
        
        current_row += len(rows) 

        # 4. Escribir nota ELD si existe
        eld_info = item.get('eld_info')
        if eld_info and eld_info.get("monto_pesos") is not None and eld_info.get("monto_pesos", 0) > 0:
            note_text = ""
            if item.get('tipo') == 'RP':
                monto_uf = eld_info.get("monto_uf", "N/A")
                monto_pesos = eld_info.get("monto_pesos", 0)
                p_uf = eld_info.get("pension_resultante_uf", "N/A")
                p_uf_float = clean_number(p_uf)
                p_pesos_calculados = p_uf_float * header_data.get('valor_uf_float', 0.0)
                note_text = f"Opcion ELD: Monto {monto_uf} UF (${monto_pesos:,.0f}) con Pension resultante: {p_uf} UF (${p_pesos_calculados:,.0f})"
            
            elif item.get('tipo') in ['RV', 'RV_Aumentada', 'REF', 'RT']:
                compania = eld_info.get("compania", "N/A")
                monto = eld_info.get("monto_pesos", 0)
                note_text = f"Mejor Oferta ELD: {compania} con ${monto:,.0f} pesos."

            current_row += 1 
            ws[f'A{current_row}'] = note_text
            ws[f'A{current_row}'].font = style_bold_text
            current_row += 1

        current_row += 2 

    # --- 5. Auto-ajustar columnas ---
    for col_letter, max_length in column_widths.items():
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col_letter].width = min(adjusted_width, 40) 

    wb.save(output)
    return output.getvalue()
